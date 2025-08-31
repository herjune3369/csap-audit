#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSAP 기술진단 Excel 리포트 생성기
Linux 진단 결과 JSON을 CSAP 공식 양식에 맞춘 Excel로 변환
"""

import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from datetime import datetime
import logging

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class CSAPExcelReportGenerator:
    """CSAP Excel 리포트 생성기"""
    
    def __init__(self):
        self.output_dir = "output"
        self.ensure_output_dir()
        
        # 색상 정의
        self.colors = {
            'good': 'C6EFCE',      # 양호: 연한 초록
            'vulnerable': 'FFC7CE', # 취약: 연한 빨강
            'info': 'FFEB9C',       # 정보: 연한 노랑
            'header': '4472C4',     # 헤더: 파랑
            'border': '000000'      # 테두리: 검정
        }
        
        # CCE 분류 매핑 (Linux 36개 항목)
        self.cce_classification = {
            # 계정 관리 (5개)
            'CCE-0001': 'U1 계정관리', 'CCE-0002': 'U1 계정관리', 'CCE-0003': 'U1 계정관리',
            'CCE-0004': 'U1 계정관리', 'CCE-0005': 'U1 계정관리',
            
            # 파일 및 디렉터리 관리 (14개)
            'CCE-0006': 'U2 파일관리', 'CCE-0007': 'U2 파일관리', 'CCE-0008': 'U2 파일관리',
            'CCE-0009': 'U2 파일관리', 'CCE-0010': 'U2 파일관리', 'CCE-0011': 'U2 파일관리',
            'CCE-0012': 'U2 파일관리', 'CCE-0013': 'U2 파일관리', 'CCE-0014': 'U2 파일관리',
            'CCE-0015': 'U2 파일관리', 'CCE-0016': 'U2 파일관리', 'CCE-0017': 'U2 파일관리',
            'CCE-0018': 'U2 파일관리', 'CCE-0019': 'U2 파일관리',
            
            # 서비스 관리 (15개)
            'CCE-0020': 'U3 서비스관리', 'CCE-0021': 'U3 서비스관리', 'CCE-0022': 'U3 서비스관리',
            'CCE-0023': 'U3 서비스관리', 'CCE-0024': 'U3 서비스관리', 'CCE-0025': 'U3 서비스관리',
            'CCE-0026': 'U3 서비스관리', 'CCE-0027': 'U3 서비스관리', 'CCE-0028': 'U3 서비스관리',
            'CCE-0029': 'U3 서비스관리', 'CCE-0030': 'U3 서비스관리', 'CCE-0031': 'U3 서비스관리',
            'CCE-0032': 'U3 서비스관리', 'CCE-0033': 'U3 서비스관리', 'CCE-0034': 'U3 서비스관리',
            
            # 패치 및 로그 관리 (2개)
            'CCE-0035': 'U4 패치관리', 'CCE-0036': 'U4 패치관리'
        }
        
        # 중요도 매핑 (기본값: H)
        self.importance_mapping = {
            # 계정 관리 - 모두 High
            'CCE-0001': 'H', 'CCE-0002': 'H', 'CCE-0003': 'H', 'CCE-0004': 'H', 'CCE-0005': 'H',
            
            # 파일 관리 - 모두 High
            'CCE-0006': 'H', 'CCE-0007': 'H', 'CCE-0008': 'H', 'CCE-0009': 'H', 'CCE-0010': 'H',
            'CCE-0011': 'H', 'CCE-0012': 'H', 'CCE-0013': 'H', 'CCE-0014': 'H', 'CCE-0015': 'H',
            'CCE-0016': 'H', 'CCE-0017': 'H', 'CCE-0018': 'H', 'CCE-0019': 'H',
            
            # 서비스 관리 - 모두 High
            'CCE-0020': 'H', 'CCE-0021': 'H', 'CCE-0022': 'H', 'CCE-0023': 'H', 'CCE-0024': 'H',
            'CCE-0025': 'H', 'CCE-0026': 'H', 'CCE-0027': 'H', 'CCE-0028': 'H', 'CCE-0029': 'H',
            'CCE-0030': 'H', 'CCE-0031': 'H', 'CCE-0032': 'H', 'CCE-0033': 'H', 'CCE-0034': 'H',
            
            # 패치 관리 - 모두 High
            'CCE-0035': 'H', 'CCE-0036': 'H'
        }
    
    def ensure_output_dir(self):
        """출력 디렉토리 생성"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"출력 디렉토리 생성: {self.output_dir}")
    
    def load_json_data(self, json_file_path):
        """JSON 파일 로드"""
        try:
            with open(json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            logger.info(f"JSON 파일 로드 완료: {json_file_path}")
            return data
        except Exception as e:
            logger.error(f"JSON 파일 로드 실패: {e}")
            raise
    
    def transform_json_to_csap_format(self, json_data):
        """JSON 데이터를 CSAP 형식으로 변환"""
        results = []
        
        if 'results' in json_data:
            for item in json_data['results']:
                cce_id = item.get('CCE_ID', '')
                classification = self.cce_classification.get(cce_id, 'U0 기타')
                importance = self.importance_mapping.get(cce_id, 'H')
                
                result = {
                    'CCE_ID': cce_id,
                    '분류': classification,
                    '점검항목': item.get('항목', ''),
                    '중요도': importance,
                    '결과': item.get('결과', ''),
                    '현황': item.get('detail', ''),
                    '개선방안': item.get('remediation', '')
                }
                results.append(result)
        
        return results
    
    def generate_summary_sheet(self, json_data):
        """장비별 점검 요약 시트 생성"""
        results = self.transform_json_to_csap_format(json_data)
        
        # 분류별 통계 계산
        summary_data = {}
        for item in results:
            classification = item['분류']
            result = item['결과']
            
            if classification not in summary_data:
                summary_data[classification] = {'전체': 0, '양호': 0, '취약': 0, '정보': 0}
            
            summary_data[classification]['전체'] += 1
            
            if result == '양호':
                summary_data[classification]['양호'] += 1
            elif result == '취약':
                summary_data[classification]['취약'] += 1
            elif result == '정보':
                summary_data[classification]['정보'] += 1
        
        # DataFrame 생성
        summary_rows = []
        for classification, stats in summary_data.items():
            total = stats['전체']
            good = stats['양호']
            vulnerable = stats['취약']
            info = stats['정보']
            security_level = round((good / total) * 100, 1) if total > 0 else 0
            
            summary_rows.append({
                '분류': classification,
                '전체 항목 수': total,
                '양호 수': good,
                '취약 수': vulnerable,
                '정보 수': info,
                '보안수준(%)': security_level
            })
        
        return pd.DataFrame(summary_rows)
    
    def generate_detail_sheet(self, json_data):
        """상세 점검 결과 시트 생성"""
        results = self.transform_json_to_csap_format(json_data)
        
        # DataFrame 생성
        detail_rows = []
        for i, item in enumerate(results, 1):
            detail_rows.append({
                'No': i,
                '분류': item['분류'],
                'CCE ID': item['CCE_ID'],
                '점검 항목': item['점검항목'],
                '중요도': item['중요도'],
                '결과': item['결과'],
                '현황': item['현황'],
                '개선방안': item['개선방안']
            })
        
        return pd.DataFrame(detail_rows)
    
    def apply_excel_styling(self, workbook):
        """Excel 스타일 적용"""
        # 기본 폰트 설정
        default_font = Font(name='맑은 고딕', size=10)
        header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
        
        # 색상 정의
        header_fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        good_fill = PatternFill(start_color=self.colors['good'], end_color=self.colors['good'], fill_type='solid')
        vulnerable_fill = PatternFill(start_color=self.colors['vulnerable'], end_color=self.colors['vulnerable'], fill_type='solid')
        info_fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
        
        # 테두리 스타일
        thin_border = Border(
            left=Side(style='thin', color=self.colors['border']),
            right=Side(style='thin', color=self.colors['border']),
            top=Side(style='thin', color=self.colors['border']),
            bottom=Side(style='thin', color=self.colors['border'])
        )
        
        # 각 워크시트에 스타일 적용
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # 헤더 스타일 적용
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 데이터 행 스타일 적용
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = default_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # 결과 열에 따른 색상 적용
                if sheet_name == '상세결과':
                    result_cell = row[4]  # 결과 열 (E열)
                    if result_cell.value == '양호':
                        result_cell.fill = good_fill
                    elif result_cell.value == '취약':
                        result_cell.fill = vulnerable_fill
                    elif result_cell.value == '정보':
                        result_cell.fill = info_fill
            
            # 열 너비 자동 조정
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def save_excel_report(self, json_file_path, output_file_path=None):
        """Excel 리포트 생성 및 저장"""
        try:
            # JSON 데이터 로드
            json_data = self.load_json_data(json_file_path)
            
            # Excel 워크북 생성
            workbook = openpyxl.Workbook()
            
            # 기본 시트 제거
            workbook.remove(workbook.active)
            
            # 요약 시트 생성
            summary_df = self.generate_summary_sheet(json_data)
            summary_sheet = workbook.create_sheet("요약")
            
            # 요약 데이터 추가
            for r in dataframe_to_rows(summary_df, index=False, header=True):
                summary_sheet.append(r)
            
            # 상세 결과 시트 생성
            detail_df = self.generate_detail_sheet(json_data)
            detail_sheet = workbook.create_sheet("상세결과")
            
            # 상세 데이터 추가
            for r in dataframe_to_rows(detail_df, index=False, header=True):
                detail_sheet.append(r)
            
            # 스타일 적용
            self.apply_excel_styling(workbook)
            
            # 파일 저장
            if output_file_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file_path = os.path.join(self.output_dir, f"linux_report_{timestamp}.xlsx")
            
            workbook.save(output_file_path)
            logger.info(f"Excel 리포트 생성 완료: {output_file_path}")
            
            return output_file_path
            
        except Exception as e:
            logger.error(f"Excel 리포트 생성 실패: {e}")
            raise

def main():
    """메인 실행 함수"""
    generator = CSAPExcelReportGenerator()
    
    # 입력 파일 경로들
    input_files = [
        "/tmp/linux_result.json",
        "/output/linux_result.json",
        "example_output.json",
        "test/sample_linux_result.json",
        "../example_output.json"
    ]
    
    # 사용 가능한 입력 파일 찾기
    input_file = None
    for file_path in input_files:
        if os.path.exists(file_path):
            input_file = file_path
            break
    
    if input_file is None:
        logger.error("입력 JSON 파일을 찾을 수 없습니다.")
        print("사용 가능한 파일:")
        for file_path in input_files:
            if os.path.exists(file_path):
                print(f"  ✅ {file_path}")
            else:
                print(f"  ❌ {file_path}")
        return
    
    try:
        # Excel 리포트 생성
        output_file = generator.save_excel_report(input_file)
        print(f"✅ Excel 리포트 생성 완료: {output_file}")
        
        # 생성된 파일 정보 출력
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file) / 1024  # KB
            print(f"📄 파일 크기: {file_size:.1f} KB")
            
            # 워크북 정보 출력
            workbook = openpyxl.load_workbook(output_file)
            print(f"📋 시트 수: {len(workbook.sheetnames)}")
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                print(f"   - {sheet_name}: {worksheet.max_row}행 x {worksheet.max_column}열")
    
    except Exception as e:
        logger.error(f"리포트 생성 중 오류 발생: {e}")

if __name__ == "__main__":
    main() 