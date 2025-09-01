#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
다중 시스템 LLM 기반 Excel 보고서 자동화 시스템
Linux, DB, Web 등 다양한 시스템의 진단 결과를 LLM으로 해설하여 Excel 보고서 생성
"""

import json
import os
import logging
import asyncio
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path

# LLM 관련
import google.generativeai as genai
from dotenv import load_dotenv

# Excel 관련
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# 로깅 설정
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# .env 파일 로드
load_dotenv()


class LLMReportGenerator:
    """다중 시스템 LLM 기반 Excel 리포트 생성기"""

    def __init__(self):
        self.output_dir = "output"
        self.ensure_output_dir()

        # Gemini API 설정
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            logger.warning("GEMINI_API_KEY가 설정되지 않았습니다. LLM 기능이 제한됩니다.")
            self.llm_enabled = False
        else:
            genai.configure(api_key=api_key)
            self.model = genai.GenerativeModel("gemini-1.5-flash")
            self.llm_enabled = True

        # 색상 정의
        self.colors = {
            "header": "4472C4",  # 헤더: 파랑
            "good": "C6EFCE",  # 양호: 연한 초록
            "vulnerable": "FFC7CE",  # 취약: 연한 빨강
            "info": "FFEB9C",  # 정보: 연한 노랑
            "border": "000000",  # 테두리: 검정
            "light_gray": "F2F2F2",  # 연한 회색
        }

        # 폰트 정의
        self.fonts = {
            "header": Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF"),
            "title": Font(name="맑은 고딕", size=12, bold=True),
            "normal": Font(name="맑은 고딕", size=9),
            "bold": Font(name="맑은 고딕", size=9, bold=True),
        }

        # 정렬 정의
        self.alignments = {
            "center": Alignment(horizontal="center", vertical="center"),
            "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "right": Alignment(horizontal="right", vertical="center"),
        }

        # 테두리 정의
        self.borders = {
            "thin": Border(
                left=Side(style="thin", color=self.colors["border"]),
                right=Side(style="thin", color=self.colors["border"]),
                top=Side(style="thin", color=self.colors["border"]),
                bottom=Side(style="thin", color=self.colors["border"]),
            )
        }

    def ensure_output_dir(self):
        """출력 디렉토리 생성"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"출력 디렉토리 생성: {self.output_dir}")

    def load_json_data(self, json_file_path: str) -> List[Dict[str, Any]]:
        """JSON 파일 로드"""
        try:
            with open(json_file_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            # 기존 형식과 새로운 형식 모두 지원
            if isinstance(data, dict) and "results" in data:
                return data["results"]
            elif isinstance(data, list):
                return data
            else:
                logger.error(f"지원하지 않는 JSON 형식: {json_file_path}")
                return []

        except Exception as e:
            logger.error(f"JSON 파일 로드 실패: {e}")
            return []

    async def enhance_with_llm(self, item: Dict[str, Any]) -> Dict[str, Any]:
        """LLM을 사용하여 항목 정보를 확장"""
        if not self.llm_enabled:
            # LLM이 비활성화된 경우 기본값 반환
            return {
                **item,
                "상세해설": "LLM 기능이 비활성화되어 있습니다.",
                "조치방법": ["LLM 기능이 비활성화되어 있습니다."],
            }

        try:
            system = item.get("시스템", "Linux")
            cce_id = item.get("CCE_ID", "")
            check_item = item.get("점검항목", "")
            result = item.get("결과", "")
            status = item.get("현황", "")
            improvement = item.get("개선방안", "")

            prompt = f"""
다음은 {system} 시스템 보안 점검 결과입니다.

● CCE ID: {cce_id}
● 점검 항목: {check_item}
● 결과: {result}
● 현황: {status}
● 개선방안: {improvement}

질문:
이 항목의 목적과 보안 중요성을 설명하고, 실제 {system} 시스템 기준으로 조치하는 방법을 알려주세요.

다음 JSON 형식으로 답변해주세요:
{{
    "상세해설": "취약점에 대한 상세한 기술적 설명",
    "조치방법": ["1단계: ...", "2단계: ...", "3단계: ..."]
}}
"""

            response = await self.model.generate_content_async(prompt)
            response_text = response.text

            # JSON 응답 파싱
            try:
                llm_data = json.loads(response_text)
                return {
                    **item,
                    "상세해설": llm_data.get("상세해설", ""),
                    "조치방법": llm_data.get("조치방법", []),
                }
            except json.JSONDecodeError:
                # JSON 파싱 실패 시 텍스트를 그대로 사용
                return {**item, "상세해설": response_text, "조치방법": ["LLM 응답을 파싱할 수 없습니다."]}

        except Exception as e:
            logger.error(f"LLM 처리 중 오류: {e}")
            return {
                **item,
                "상세해설": f"LLM 처리 중 오류: {e}",
                "조치방법": ["LLM 처리 중 오류가 발생했습니다."],
            }

    async def process_system_data(self, json_file_path: str) -> List[Dict[str, Any]]:
        """시스템별 데이터를 LLM으로 처리"""
        # JSON 데이터 로드
        items = self.load_json_data(json_file_path)
        if not items:
            return []

        # LLM으로 각 항목 확장
        enhanced_items = []
        for item in items:
            enhanced_item = await self.enhance_with_llm(item)
            enhanced_items.append(enhanced_item)
            logger.info(f"LLM 처리 완료: {item.get('CCE_ID', 'Unknown')}")

        return enhanced_items

    def generate_excel_report(
        self, enhanced_items: List[Dict[str, Any]], system_name: str
    ) -> str:
        """확장된 데이터로 Excel 보고서 생성"""
        workbook = Workbook()
        workbook.remove(workbook.active)

        # 상세 점검 결과 시트 생성
        worksheet = workbook.create_sheet("상세 점검 결과")

        # 헤더 설정
        headers = [
            "No",
            "시스템",
            "분류",
            "CCE ID",
            "점검 항목",
            "중요도",
            "결과",
            "현황",
            "개선방안",
            "상세해설",
            "조치 방법",
        ]

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = self.fonts["header"]
            cell.fill = PatternFill(
                start_color=self.colors["header"],
                end_color=self.colors["header"],
                fill_type="solid",
            )
            cell.alignment = self.alignments["center"]
            cell.border = self.borders["thin"]

        # 데이터 입력
        for row, item in enumerate(enhanced_items, 2):
            worksheet.cell(row=row, column=1, value=row - 1).border = self.borders[
                "thin"
            ]  # No
            worksheet.cell(
                row=row, column=2, value=item.get("시스템", system_name)
            ).border = self.borders["thin"]
            worksheet.cell(
                row=row, column=3, value=item.get("분류", "")
            ).border = self.borders["thin"]
            worksheet.cell(
                row=row, column=4, value=item.get("CCE_ID", "")
            ).border = self.borders["thin"]
            worksheet.cell(
                row=row, column=5, value=item.get("점검항목", "")
            ).border = self.borders["thin"]

            # 중요도
            importance_cell = worksheet.cell(
                row=row, column=6, value=item.get("중요도", "")
            )
            importance_cell.border = self.borders["thin"]
            if item.get("중요도") == "H":
                importance_cell.font = self.fonts["bold"]

            # 결과 (색상 적용)
            result_cell = worksheet.cell(row=row, column=7, value=item.get("결과", ""))
            result_cell.border = self.borders["thin"]
            if item.get("결과") == "양호":
                result_cell.fill = PatternFill(
                    start_color=self.colors["good"],
                    end_color=self.colors["good"],
                    fill_type="solid",
                )
            elif item.get("결과") == "취약":
                result_cell.fill = PatternFill(
                    start_color=self.colors["vulnerable"],
                    end_color=self.colors["vulnerable"],
                    fill_type="solid",
                )
            elif item.get("결과") == "정보":
                result_cell.fill = PatternFill(
                    start_color=self.colors["info"],
                    end_color=self.colors["info"],
                    fill_type="solid",
                )

            worksheet.cell(
                row=row, column=8, value=item.get("현황", "")
            ).border = self.borders["thin"]
            worksheet.cell(
                row=row, column=9, value=item.get("개선방안", "")
            ).border = self.borders["thin"]
            worksheet.cell(
                row=row, column=10, value=item.get("상세해설", "")
            ).border = self.borders["thin"]

            # 조치 방법 (줄바꿈으로 구분)
            action_methods = item.get("조치방법", [])
            if isinstance(action_methods, list):
                action_text = "\n".join(action_methods)
            else:
                action_text = str(action_methods)
            worksheet.cell(row=row, column=11, value=action_text).border = self.borders[
                "thin"
            ]

        # 열 너비 자동 조정
        column_widths = [5, 10, 15, 12, 25, 8, 8, 20, 20, 30, 40]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=col).column_letter
            ].width = width

        # 필터 추가
        worksheet.auto_filter.ref = worksheet.dimensions

        # 요약 통계 시트 생성
        self.generate_summary_sheet(workbook, enhanced_items, system_name)

        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"csap_{system_name.lower()}_report_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, output_filename)

        workbook.save(output_path)
        logger.info(f"Excel 리포트 생성 완료: {output_path}")

        return output_path

    def generate_summary_sheet(
        self, workbook: Workbook, items: List[Dict[str, Any]], system_name: str
    ):
        """요약 통계 시트 생성"""
        worksheet = workbook.create_sheet("요약 통계")

        # 분류별 통계 계산
        summary_data = {}
        for item in items:
            classification = item.get("분류", "기타")
            result = item.get("결과", "")

            if classification not in summary_data:
                summary_data[classification] = {"전체": 0, "양호": 0, "취약": 0, "정보": 0}

            summary_data[classification]["전체"] += 1

            if result == "양호":
                summary_data[classification]["양호"] += 1
            elif result == "취약":
                summary_data[classification]["취약"] += 1
            elif result == "정보":
                summary_data[classification]["정보"] += 1

        # 헤더 설정
        headers = ["분류", "전체 항목 수", "양호 수", "취약 수", "정보 수", "보안수준(%)"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = self.fonts["header"]
            cell.fill = PatternFill(
                start_color=self.colors["header"],
                end_color=self.colors["header"],
                fill_type="solid",
            )
            cell.alignment = self.alignments["center"]
            cell.border = self.borders["thin"]

        # 데이터 입력
        row = 2
        for classification, stats in summary_data.items():
            total = stats["전체"]
            good = stats["양호"]
            vulnerable = stats["취약"]
            info = stats["정보"]
            security_level = round((good / total) * 100, 1) if total > 0 else 0

            worksheet.cell(
                row=row, column=1, value=classification
            ).border = self.borders["thin"]
            worksheet.cell(row=row, column=2, value=total).border = self.borders["thin"]
            worksheet.cell(row=row, column=3, value=good).border = self.borders["thin"]
            worksheet.cell(row=row, column=4, value=vulnerable).border = self.borders[
                "thin"
            ]
            worksheet.cell(row=row, column=5, value=info).border = self.borders["thin"]
            worksheet.cell(
                row=row, column=6, value=security_level
            ).border = self.borders["thin"]

            # 보안수준에 따른 색상 적용
            if security_level >= 80:
                worksheet.cell(row=row, column=6).fill = PatternFill(
                    start_color=self.colors["good"],
                    end_color=self.colors["good"],
                    fill_type="solid",
                )
            elif security_level >= 60:
                worksheet.cell(row=row, column=6).fill = PatternFill(
                    start_color=self.colors["info"],
                    end_color=self.colors["info"],
                    fill_type="solid",
                )
            else:
                worksheet.cell(row=row, column=6).fill = PatternFill(
                    start_color=self.colors["vulnerable"],
                    end_color=self.colors["vulnerable"],
                    fill_type="solid",
                )

            row += 1

        # 열 너비 자동 조정
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    async def process_system_file(self, json_file_path: str) -> str:
        """시스템별 JSON 파일을 처리하여 Excel 보고서 생성"""
        try:
            # 파일명에서 시스템명 추출
            filename = os.path.basename(json_file_path)
            system_name = filename.split("_")[0] if "_" in filename else "unknown"

            logger.info(f"시스템 처리 시작: {system_name}")

            # LLM으로 데이터 확장
            enhanced_items = await self.process_system_data(json_file_path)

            if not enhanced_items:
                logger.error(f"처리할 데이터가 없습니다: {json_file_path}")
                return ""

            # Excel 보고서 생성
            output_path = self.generate_excel_report(enhanced_items, system_name)

            logger.info(f"시스템 처리 완료: {system_name} -> {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"시스템 처리 중 오류: {e}")
            return ""


async def main():
    """메인 실행 함수"""
    generator = LLMReportGenerator()

    # 입력 파일 경로들
    input_files = [
        "output/linux_result.json",
        "output/web_result.json",
        "output/db_result.json",
        "example_output.json",
    ]

    # 사용 가능한 입력 파일 찾기
    available_files = []
    for file_path in input_files:
        if os.path.exists(file_path):
            available_files.append(file_path)

    if not available_files:
        logger.error("처리할 JSON 파일을 찾을 수 없습니다.")
        return

    logger.info(f"처리할 파일들: {available_files}")

    # 각 시스템별로 처리
    for json_file in available_files:
        try:
            output_file = await generator.process_system_file(json_file)
            if output_file:
                print(f"✅ Excel 리포트 생성 완료: {output_file}")

                # 생성된 파일 정보 출력
                if os.path.exists(output_file):
                    file_size = os.path.getsize(output_file) / 1024  # KB
                    print(f"📄 파일 크기: {file_size:.1f} KB")

        except Exception as e:
            logger.error(f"파일 처리 중 오류: {json_file} - {e}")


if __name__ == "__main__":
    asyncio.run(main())
