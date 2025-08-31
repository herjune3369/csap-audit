#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 보고서 생성 테스트 스크립트
모의 LLM 응답을 사용해서 CSAP Excel 보고서 생성
"""

import json
import sys
import os
from pathlib import Path

# 상위 디렉토리의 report_generator 모듈 import
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'report_generator'))
from generate_csap_excel import CSAPExcelReportGenerator

def test_excel_generation():
    """Excel 보고서 생성 테스트"""
    try:
        # 모의 LLM 응답 파일 로드
        mock_responses_file = "output/mock_llm_responses.json"
        
        if not Path(mock_responses_file).exists():
            print(f"❌ 모의 LLM 응답 파일을 찾을 수 없습니다: {mock_responses_file}")
            return
        
        with open(mock_responses_file, 'r', encoding='utf-8') as f:
            mock_results = json.load(f)
        
        print(f"📋 {len(mock_results)}개 모의 LLM 응답 로드 완료")
        
        # LLM 응답을 원본 데이터와 통합
        enhanced_items = []
        
        for result in mock_results:
            original_item = result['original_item']
            llm_response = result['llm_response']
            
            if llm_response['success']:
                # LLM 응답 데이터 추가
                enhanced_item = {
                    **original_item,
                    '시스템': 'Linux',
                    '상세해설': llm_response['data'].get('상세해설', ''),
                    '공격시나리오': llm_response['data'].get('공격시나리오', ''),
                    '비전문가설명': llm_response['data'].get('비전문가설명', ''),
                    '조치방법': llm_response['data'].get('조치방법', [])
                }
            else:
                # LLM 실패 시 기본값
                enhanced_item = {
                    **original_item,
                    '시스템': 'Linux',
                    '상세해설': 'LLM 처리 실패',
                    '공격시나리오': 'LLM 처리 실패',
                    '비전문가설명': 'LLM 처리 실패',
                    '조치방법': ['LLM 처리 실패']
                }
            
            enhanced_items.append(enhanced_item)
        
        print(f"✅ {len(enhanced_items)}개 항목 통합 완료")
        
        # Excel 보고서 생성
        generator = CSAPExcelReportGenerator()
        
        # 임시 JSON 파일 생성
        temp_json_file = "output/temp_enhanced_items.json"
        with open(temp_json_file, 'w', encoding='utf-8') as f:
            json.dump({
                "results": enhanced_items
            }, f, ensure_ascii=False, indent=2)
        
        output_path = generator.generate_csap_excel(temp_json_file)
        
        print(f"🎉 Excel 보고서 생성 완료: {output_path}")
        
        # 생성된 파일 정보 출력
        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path) / 1024  # KB
            print(f"📄 파일 크기: {file_size:.1f} KB")
            
            # 워크북 정보 출력
            from openpyxl import load_workbook
            workbook = load_workbook(output_path)
            print(f"📋 시트 수: {len(workbook.sheetnames)}")
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                print(f"   - {sheet_name}: {worksheet.max_row}행 x {worksheet.max_column}열")
        
    except Exception as e:
        print(f"❌ Excel 보고서 생성 중 오류: {e}")

if __name__ == "__main__":
    test_excel_generation() 