#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSAP 기술진단 Excel 리포트 자동 생성기
CSAP PDF 양식과 동일한 구조로 Excel 보고서 생성
"""

import json
import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from typing import List, Dict, Any

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class CSAPExcelReportGenerator:
    """CSAP Excel 리포트 생성기 (PDF 양식 기반)"""
    
    def __init__(self):
        self.output_dir = "output"
        self.ensure_output_dir()
        
        # 색상 정의 (CSAP 표준)
        self.colors = {
            'header': '4472C4',      # 헤더: 파랑
            'good': 'C6EFCE',        # 양호: 연한 초록
            'vulnerable': 'FFC7CE',   # 취약: 연한 빨강
            'info': 'FFEB9C',        # 정보: 연한 노랑
            'border': '000000',       # 테두리: 검정
            'light_gray': 'F2F2F2'   # 연한 회색
        }
        
        # 폰트 정의
        self.fonts = {
            'header': Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF'),
            'title': Font(name='맑은 고딕', size=14, bold=True),
            'normal': Font(name='맑은 고딕', size=10),
            'bold': Font(name='맑은 고딕', size=10, bold=True)
        }
        
        # 정렬 정의
        self.alignments = {
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center')
        }
        
        # 테두리 정의
        self.borders = {
            'thin': Border(
                left=Side(style='thin', color=self.colors['border']),
                right=Side(style='thin', color=self.colors['border']),
                top=Side(style='thin', color=self.colors['border']),
                bottom=Side(style='thin', color=self.colors['border'])
            )
        }
    
    def ensure_output_dir(self):
        """출력 디렉토리 생성"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"출력 디렉토리 생성: {self.output_dir}")
    
    def load_json_data(self, json_file_path):
        """JSON 파일 로드"""
        try:
            with open(json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            logger.info(f"JSON 파일 로드 완료: {json_file_path}")
            return data
        except Exception as e:
            logger.error(f"JSON 파일 로드 실패: {e}")
            raise
    
    def transform_json_to_csap_format(self, json_data):
        """JSON 데이터를 CSAP 형식으로 변환 (LLM 데이터 포함)"""
        results = []
        
        if 'results' in json_data:
            for item in json_data['results']:
                cce_id = item.get('CCE_ID', '')
                classification = self.get_classification(cce_id)
                importance = self.get_importance(cce_id)
                
                result = {
                    'CCE_ID': cce_id,
                    '분류': classification,
                    '점검항목': item.get('항목', item.get('점검항목', '')),
                    '중요도': importance,
                    '결과': item.get('결과', ''),
                    '현황': item.get('detail', item.get('현황', '')),
                    '개선방안': item.get('remediation', item.get('개선방안', ''))
                }
                
                # LLM 상세 정보가 있는 경우 추가
                if '상세해설' in item:
                    result['상세해설'] = item.get('상세해설', '')
                    result['공격시나리오'] = item.get('공격시나리오', '')
                    result['비전문가설명'] = item.get('비전문가설명', '')
                    result['조치방법'] = item.get('조치방법', [])
                    result['시스템'] = item.get('시스템', '')
                
                results.append(result)
        
        return results
    
    def get_classification(self, cce_id):
        """CCE ID에 따른 분류 반환"""
        # CCE 분류 매핑 (Linux 36개 항목)
        classification_map = {
            # 계정 관리 (5개)
            'CCE-0001': 'U1 계정관리', 'CCE-0002': 'U1 계정관리', 'CCE-0003': 'U1 계정관리',
            'CCE-0004': 'U1 계정관리', 'CCE-0005': 'U1 계정관리',
            
            # 파일 및 디렉터리 관리 (14개)
            'CCE-0006': 'U2 파일관리', 'CCE-0007': 'U2 파일관리', 'CCE-0008': 'U2 파일관리',
            'CCE-0009': 'U2 파일관리', 'CCE-0010': 'U2 파일관리', 'CCE-0011': 'U2 파일관리',
            'CCE-0012': 'U2 파일관리', 'CCE-0013': 'U2 파일관리', 'CCE-0014': 'U2 파일관리',
            'CCE-0015': 'U2 파일관리', 'CCE-0016': 'U2 파일관리', 'CCE-0017': 'U2 파일관리',
            'CCE-0018': 'U2 파일관리', 'CCE-0019': 'U2 파일관리',
            
            # 서비스 관리 (15개)
            'CCE-0020': 'U3 서비스관리', 'CCE-0021': 'U3 서비스관리', 'CCE-0022': 'U3 서비스관리',
            'CCE-0023': 'U3 서비스관리', 'CCE-0024': 'U3 서비스관리', 'CCE-0025': 'U3 서비스관리',
            'CCE-0026': 'U3 서비스관리', 'CCE-0027': 'U3 서비스관리', 'CCE-0028': 'U3 서비스관리',
            'CCE-0029': 'U3 서비스관리', 'CCE-0030': 'U3 서비스관리', 'CCE-0031': 'U3 서비스관리',
            'CCE-0032': 'U3 서비스관리', 'CCE-0033': 'U3 서비스관리', 'CCE-0034': 'U3 서비스관리',
            
            # 패치 및 로그 관리 (2개)
            'CCE-0035': 'U4 패치관리', 'CCE-0036': 'U4 패치관리'
        }
        
        return classification_map.get(cce_id, 'U0 기타')
    
    def get_importance(self, cce_id):
        """CCE ID에 따른 중요도 반환"""
        # 모든 Linux 진단 항목은 High 중요도
        return 'H'
    
    def generate_summary_sheet(self, workbook, results):
        """요약 통계 시트 생성"""
        worksheet = workbook.create_sheet("요약 통계")
        
        # 분류별 통계 계산
        summary_data = {}
        for item in results:
            classification = item['분류']
            result = item['결과']
            
            if classification not in summary_data:
                summary_data[classification] = {'전체': 0, '양호': 0, '취약': 0, '정보': 0}
            
            summary_data[classification]['전체'] += 1
            
            if result == '양호':
                summary_data[classification]['양호'] += 1
            elif result == '취약':
                summary_data[classification]['취약'] += 1
            elif result == '정보':
                summary_data[classification]['정보'] += 1
        
        # 헤더 설정
        headers = ['분류', '전체 항목 수', '양호 수', '취약 수', '정보 수', '보안수준(%)']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = self.alignments['center']
            cell.border = self.borders['thin']
        
        # 데이터 입력
        row = 2
        for classification, stats in summary_data.items():
            total = stats['전체']
            good = stats['양호']
            vulnerable = stats['취약']
            info = stats['정보']
            security_level = round((good / total) * 100, 1) if total > 0 else 0
            
            worksheet.cell(row=row, column=1, value=classification).border = self.borders['thin']
            worksheet.cell(row=row, column=2, value=total).border = self.borders['thin']
            worksheet.cell(row=row, column=3, value=good).border = self.borders['thin']
            worksheet.cell(row=row, column=4, value=vulnerable).border = self.borders['thin']
            worksheet.cell(row=row, column=5, value=info).border = self.borders['thin']
            worksheet.cell(row=row, column=6, value=security_level).border = self.borders['thin']
            
            # 보안수준에 따른 색상 적용
            if security_level >= 80:
                worksheet.cell(row=row, column=6).fill = PatternFill(start_color=self.colors['good'], end_color=self.colors['good'], fill_type='solid')
            elif security_level >= 60:
                worksheet.cell(row=row, column=6).fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
            else:
                worksheet.cell(row=row, column=6).fill = PatternFill(start_color=self.colors['vulnerable'], end_color=self.colors['vulnerable'], fill_type='solid')
            
            row += 1
        
        # 열 너비 자동 조정
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 차트 추가
        self.add_summary_chart(worksheet, summary_data)
        
        return worksheet
    
    def add_summary_chart(self, worksheet, summary_data):
        """요약 차트 추가"""
        # 차트 데이터 준비
        categories = list(summary_data.keys())
        good_values = [summary_data[cat]['양호'] for cat in categories]
        vulnerable_values = [summary_data[cat]['취약'] for cat in categories]
        
        # 차트 생성
        chart = BarChart()
        chart.title = "분류별 보안 진단 결과"
        chart.x_axis.title = "분류"
        chart.y_axis.title = "항목 수"
        
        # 데이터 추가
        data = Reference(worksheet, min_col=3, min_row=1, max_row=len(categories)+1, max_col=4)
        cats = Reference(worksheet, min_col=1, min_row=2, max_row=len(categories)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # 차트 위치 설정
        worksheet.add_chart(chart, "H2")
    
    def generate_detail_sheet(self, workbook, results):
        """상세 점검 결과 시트 생성 (LLM 상세 정보 포함)"""
        worksheet = workbook.create_sheet("상세 점검 결과")
        
        # LLM 상세 정보가 있는지 확인
        has_llm_data = any('상세해설' in item for item in results)
        
        if has_llm_data:
            # LLM 상세 정보가 있는 경우 확장된 헤더
            headers = ['No', '시스템', '분류', 'CCE ID', '점검 항목', '중요도', '결과', '현황', '개선방안', '상세해설', '공격시나리오', '비전문가설명', '조치방법']
        else:
            # 기존 헤더
            headers = ['No', '분류', 'CCE ID', '점검 항목', '중요도', '결과', '현황', '개선방안']
        
        # 헤더 설정
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = self.alignments['center']
            cell.border = self.borders['thin']
        
        # 데이터 입력
        for row, item in enumerate(results, 2):
            col = 1
            worksheet.cell(row=row, column=col, value=row-1).border = self.borders['thin']  # No
            col += 1
            
            if has_llm_data:
                # 시스템 정보 추가
                worksheet.cell(row=row, column=col, value=item.get('시스템', '')).border = self.borders['thin']
                col += 1
            
            worksheet.cell(row=row, column=col, value=item['분류']).border = self.borders['thin']
            col += 1
            worksheet.cell(row=row, column=col, value=item['CCE_ID']).border = self.borders['thin']
            col += 1
            worksheet.cell(row=row, column=col, value=item['점검항목']).border = self.borders['thin']
            col += 1
            
            # 중요도 (H는 굵게)
            importance_cell = worksheet.cell(row=row, column=col, value=item['중요도'])
            importance_cell.border = self.borders['thin']
            if item['중요도'] == 'H':
                importance_cell.font = self.fonts['bold']
            col += 1
            
            # 결과 (색상 적용)
            result_cell = worksheet.cell(row=row, column=col, value=item['결과'])
            result_cell.border = self.borders['thin']
            if item['결과'] == '양호':
                result_cell.fill = PatternFill(start_color=self.colors['good'], end_color=self.colors['good'], fill_type='solid')
            elif item['결과'] == '취약':
                result_cell.fill = PatternFill(start_color=self.colors['vulnerable'], end_color=self.colors['vulnerable'], fill_type='solid')
            elif item['결과'] == '정보':
                result_cell.fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
            col += 1
            
            worksheet.cell(row=row, column=col, value=item['현황']).border = self.borders['thin']
            col += 1
            worksheet.cell(row=row, column=col, value=item['개선방안']).border = self.borders['thin']
            col += 1
            
            if has_llm_data:
                # LLM 상세 정보 추가
                worksheet.cell(row=row, column=col, value=item.get('상세해설', '')).border = self.borders['thin']
                col += 1
                worksheet.cell(row=row, column=col, value=item.get('공격시나리오', '')).border = self.borders['thin']
                col += 1
                worksheet.cell(row=row, column=col, value=item.get('비전문가설명', '')).border = self.borders['thin']
                col += 1
                
                # 조치방법 (리스트를 문자열로 변환)
                action_methods = item.get('조치방법', [])
                if isinstance(action_methods, list):
                    action_text = '\n'.join(action_methods)
                else:
                    action_text = str(action_methods)
                worksheet.cell(row=row, column=col, value=action_text).border = self.borders['thin']
                worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 열 너비 자동 조정
        if has_llm_data:
            column_widths = [5, 10, 15, 12, 25, 8, 8, 30, 30, 40, 40, 40, 50]
        else:
            column_widths = [5, 15, 12, 25, 8, 8, 30, 30]
        
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = width
        
        # 필터 추가
        worksheet.auto_filter.ref = worksheet.dimensions
        
        return worksheet
    
    def generate_csap_excel(self, json_path, output_path=None):
        """CSAP Excel 리포트 생성"""
        try:
            # JSON 데이터 로드
            json_data = self.load_json_data(json_path)
            
            # Excel 워크북 생성
            workbook = Workbook()
            
            # 기본 시트 제거
            workbook.remove(workbook.active)
            
            # JSON 데이터를 CSAP 형식으로 변환
            results = self.transform_json_to_csap_format(json_data)
            
            # 요약 통계 시트 생성
            self.generate_summary_sheet(workbook, results)
            
            # 상세 점검 결과 시트 생성
            self.generate_detail_sheet(workbook, results)
            
            # 파일 저장
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = os.path.join(self.output_dir, f"csap_linux_report_{timestamp}.xlsx")
            
            workbook.save(output_path)
            logger.info(f"CSAP Excel 리포트 생성 완료: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"CSAP Excel 리포트 생성 실패: {e}")
            raise

    def generate_csap_excel_from_items(self, enhanced_items: List[Dict[str, Any]], system_name: str) -> str:
        """확장된 항목 리스트로 Excel 보고서 생성"""
        try:
            # Excel 보고서 생성
            output_path = self.generate_excel_report(enhanced_items, system_name)
            
            logger.info(f"CSAP Excel 리포트 생성 완료: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"CSAP Excel 리포트 생성 실패: {e}")
            raise

def main():
    """메인 실행 함수"""
    generator = CSAPExcelReportGenerator()
    
    # 입력 파일 경로들
    input_files = [
        "/tmp/linux_result.json",
        "/output/linux_result.json",
        "example_output.json",
        "test/sample_linux_result.json"
    ]
    
    # 사용 가능한 입력 파일 찾기
    input_file = None
    for file_path in input_files:
        if os.path.exists(file_path):
            input_file = file_path
            break
    
    if input_file is None:
        logger.error("입력 JSON 파일을 찾을 수 없습니다.")
        return
    
    try:
        # CSAP Excel 리포트 생성
        output_file = generator.generate_csap_excel(input_file)
        print(f"✅ CSAP Excel 리포트 생성 완료: {output_file}")
        
        # 생성된 파일 정보 출력
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file) / 1024  # KB
            print(f"📄 파일 크기: {file_size:.1f} KB")
            
            # 워크북 정보 출력
            workbook = load_workbook(output_file)
            print(f"📋 시트 수: {len(workbook.sheetnames)}")
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                print(f"   - {sheet_name}: {worksheet.max_row}행 x {worksheet.max_column}열")
    
    except Exception as e:
        logger.error(f"리포트 생성 중 오류 발생: {e}")

if __name__ == "__main__":
    main() 